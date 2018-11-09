from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from typing import List, Iterator


class KaWorkBook(Workbook):
    def write_data(self, data: List[List[str]], sheetname: str = 'Sheet') -> None:
        s = self.create_sheet(sheetname)
        for row, _data in enumerate(data):
            for column, __data in enumerate(_data):
                s.cell(row=row + 1, column=column + 1).value = __data


class OpenExcel(object):

    def __init__(self, filename, methods, sheetname='Sheet1'):
        self.filename = filename
        self.methods = methods
        self.m = self.enter_methods()
        self.sheetname = sheetname

    def write_cursor(self):
        wb = KaWorkBook()
        return wb

    def read_cursor(self):
        wb = load_workbook(self.filename)
        return wb

    def enter_methods(self):
        d = {
            'w': self.write_cursor,
            'r': self.read_cursor,
        }
        if self.methods not in d.keys():
            raise KeyboardInterrupt
        return d.get(self.methods, None)()

    def save_data(self, data: List[Iterator[str]]) -> None:
        if isinstance(self.m, self.write_cursor):
            s = self.m.create_sheet(self.sheetname)
            for row, _data in enumerate(data):
                for column, __data in enumerate(_data):
                    s.cell(row=row + 1, column=column + 1).value = __data
        else:
            return

    def exit_methods(self):
        d = {
            'w': self.m.save(self.filename),
            'r': self.m.close(),
        }
        return d.get(self.methods, None)
    
    def __enter__(self):
        return self.m

    def __exit__(self, exc_type, exc_value, exc_tb):
        self.exit_methods()
